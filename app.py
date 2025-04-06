import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook  # For Excel file handling
from PIL import Image, ImageTk  # For adding images
import os  # To scan for .db files
import sys

# Helper function to handle file paths for PyInstaller
def resource_path(relative_path):
    """Get the absolute path to a resource, works for dev and PyInstaller."""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

# Global variables
current_db = None
conn = None
first_major = None  # To store the major of the first student who taps their tag
first_stage = None  # To store the stage of the first student
first_study = None  # To store the study of the first student
first_group = None  # To store the group of the first student

def create_db_connection(db_file):
    """Create a connection to the SQLite database."""
    global conn
    try:
        conn = sqlite3.connect(db_file)
        return True
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", str(e))
        return False

def initialize_db():
    """Initialize the database with the required tables."""
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                major TEXT NOT NULL,
                stage TEXT NOT NULL,
                study TEXT NOT NULL,
                group_name TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT,
                name TEXT,
                major TEXT,
                stage TEXT,
                study TEXT,
                group_name TEXT,
                timestamp TEXT,
                attended INTEGER,
                FOREIGN KEY(student_id) REFERENCES students(student_id)
            )
        """)
        conn.commit()

def load_db_from_dropdown(event=None):
    """Load a database selected from the dropdown."""
    global current_db, first_major, first_stage, first_study, first_group
    selected_db = db_dropdown.get().strip()
    if not selected_db:
        messagebox.showwarning("Input Error", "No database selected.")
        return
    db_file = os.path.join("Students Databases", selected_db)
    if create_db_connection(db_file):
        current_db = db_file
        initialize_db()
        first_major = None  # Reset filters when loading a new database
        first_stage = None
        first_study = None
        first_group = None
        messagebox.showinfo("Success", f"Database loaded: {db_file}")
    else:
        messagebox.showerror("Error", "Failed to load database.")

def populate_db_dropdown():
    """Populate the dropdown with .db files from the 'Students Databases' folder."""
    db_files = []
    try:
        # Ensure the folder exists
        if not os.path.exists("Students Databases"):
            os.makedirs("Students Databases")  # Create the folder if it doesn't exist
            messagebox.showinfo("Info", "Created 'Students Databases' folder.")
        
        db_files = [f for f in os.listdir("Students Databases") if f.endswith(".db")]
    except Exception as e:
        print(f"Error reading 'Students Databases' folder: {e}")
    db_dropdown['values'] = db_files
    if db_files:
        db_dropdown.current(0)  # Select the first database by default
    else:
        messagebox.showinfo("Information", "No database files found. Please create a new database.")

def create_db():
    """Create a new .db file."""
    global current_db, first_major, first_stage, first_study, first_group
    db_file = filedialog.asksaveasfilename(
        title="Create New Database File",
        defaultextension=".db",
        filetypes=[("SQLite Database", "*.db")]
    )
    if db_file:
        if create_db_connection(db_file):
            current_db = db_file
            initialize_db()
            first_major = None  # Reset filters when creating a new database
            first_stage = None
            first_study = None
            first_group = None
            messagebox.showinfo("Success", f"New database created: {db_file}")
            # Move the new database to the 'Students Databases' folder
            target_folder = "Students Databases"
            if not os.path.exists(target_folder):
                os.makedirs(target_folder)
            target_path = os.path.join(target_folder, os.path.basename(db_file))
            os.replace(db_file, target_path)
            populate_db_dropdown()  # Refresh the dropdown
        else:
            messagebox.showerror("Error", "Failed to create database.")

def add_student():
    """Add a new student to the database."""
    if not conn:
        messagebox.showerror("Error", "No database loaded.")
        return
    student_id = entry_id.get().strip()
    name = entry_name.get().strip()
    major = entry_major.get().strip()
    stage = entry_stage.get().strip()
    study = entry_study.get().strip()  # This will get the selected value from the combobox
    group = entry_group.get().strip()
    if not student_id or not name or not major or not stage or not study or not group:
        messagebox.showwarning("Input Error", "All fields are required.")
        return
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO students (student_id, name, major, stage, study, group_name)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (student_id, name, major, stage, study, group))
        conn.commit()
        messagebox.showinfo("Success", "Student added successfully.")
        clear_entries()
        # Reset combobox to default selection after clearing
        entry_study.current(0)
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Student ID already exists.")

def clear_entries():
    """Clear input fields."""
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_major.delete(0, tk.END)
    entry_stage.delete(0, tk.END)
    # For combobox, we don't delete, we set it back to a default value
    entry_study.current(0)  # Reset to Morning
    entry_group.delete(0, tk.END)

def import_students_from_excel():
    """Import students from an Excel file into the database."""
    if not conn:
        messagebox.showerror("Error", "No database loaded.")
        return
    excel_file = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if not excel_file:
        return  # Exit if no file is selected
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        cursor = conn.cursor()
        success_count = 0
        duplicate_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            student_id, name, major, stage, study, group = row[0], row[1], row[2], row[3], row[4], row[5]
            if not student_id or not name or not major or not stage or not study or not group:
                continue  # Skip empty rows
            try:
                cursor.execute("""
                    INSERT INTO students (student_id, name, major, stage, study, group_name)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (student_id, name, major, stage, study, group))
                success_count += 1
            except sqlite3.IntegrityError:
                duplicate_count += 1  # Count duplicates but don't stop
        conn.commit()
        message = f"Successfully imported {success_count} students.\n"
        if duplicate_count > 0:
            message += f"{duplicate_count} duplicate entries were skipped."
        messagebox.showinfo("Import Complete", message)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to import students: {str(e)}")

def record_attendance(event=None):
    """Record attendance using NFC and update the dashboard."""
    global first_major, first_stage, first_study, first_group
    if not conn:
        return  # Silently exit if no database is loaded
    nfc_id = entry_nfc.get().strip()
    if not nfc_id:
        return  # Silently exit if no NFC ID is detected
    entry_nfc.delete(0, tk.END)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students WHERE student_id=?", (nfc_id,))
    student = cursor.fetchone()
    if not student:
        return  # Silently exit if the student is not found

    # Check if the student's major, stage, study, and group match the filters
    if first_major is None:
        first_major = student[2]  # Set the first_major to the major of the first student
        first_stage = student[3]
        first_study = student[4]
        first_group = student[5]
        
        # More descriptive message when using Morning or Hosted study types
        if first_study == "Morning" or first_study == "Hosted":
            messagebox.showinfo("Info", f"Filters set to: Major={first_major}, Stage={first_stage}, Study={first_study} (will include both Morning and Hosted students), Group={first_group}")
        else:
            messagebox.showinfo("Info", f"Filters set to: Major={first_major}, Stage={first_stage}, Study={first_study}, Group={first_group}")
    elif student[2] != first_major or student[3] != first_stage or student[5] != first_group:
        # Major, stage, and group must match exactly
        messagebox.showwarning("Filter Mismatch", f"Attendance is currently restricted to students with:\nMajor={first_major}, Stage={first_stage}, Group={first_group}.")
        return
    elif not ((student[4] == first_study) or 
              (first_study == "Morning" and student[4] == "Hosted") or 
              (first_study == "Hosted" and student[4] == "Morning")):
        # Study must be either the same or a compatible type (Morning or Hosted)
        messagebox.showwarning("Filter Mismatch", f"Attendance is currently restricted to students with:\nStudy={first_study} or compatible study types.")
        return

    # Check if the student has already been marked as attended today
    today = datetime.now().strftime("%Y-%m-%d")
    cursor.execute("""
        SELECT * FROM attendance 
        WHERE student_id=? AND DATE(timestamp)=?
    """, (nfc_id, today))
    existing_record = cursor.fetchone()
    if existing_record:
        messagebox.showwarning("Already Attended", f"{student[1]} has already been marked as attended today.")
        return

    # Insert the new attendance record
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("""
        INSERT INTO attendance (student_id, name, major, stage, study, group_name, timestamp, attended)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
    """, (student[0], student[1], student[2], student[3], student[4], student[5], timestamp))
    conn.commit()

    # Update the dashboard
    dashboard_tree.insert("", tk.END, values=(student[1], student[2], student[3], student[4], student[5], timestamp, "Yes"))

def export_attendance():
    """Export attendance data to an Excel file, filtered by the first student's major, stage, study, and group."""
    global first_major, first_stage, first_study, first_group
    if not conn:
        messagebox.showerror("Error", "No database loaded.")
        return
    if first_major is None:
        messagebox.showwarning("Warning", "No attendance recorded yet. Cannot determine the filters.")
        return
    cursor = conn.cursor()
    
    # Build the where clause based on the study type
    if first_study == "Morning" or first_study == "Hosted":
        study_clause = "(s.study = 'Morning' OR s.study = 'Hosted')"
        # For Morning or Hosted, include both types in the export
        cursor.execute(f"""
            SELECT s.name, s.major, s.stage, s.study, s.group_name, a.timestamp, CASE WHEN a.attended = 1 THEN 'Yes' ELSE 'No' END AS attended
            FROM students s
            LEFT JOIN attendance a ON s.student_id = a.student_id
            WHERE s.major = ? AND s.stage = ? AND {study_clause} AND s.group_name = ?
        """, (first_major, first_stage, first_group))
    else:
        study_clause = "s.study = ?"
        cursor.execute(f"""
            SELECT s.name, s.major, s.stage, s.study, s.group_name, a.timestamp, CASE WHEN a.attended = 1 THEN 'Yes' ELSE 'No' END AS attended
            FROM students s
            LEFT JOIN attendance a ON s.student_id = a.student_id
            WHERE s.major = ? AND s.stage = ? AND {study_clause} AND s.group_name = ?
        """, (first_major, first_stage, first_study, first_group))
    
    rows = cursor.fetchall()
    if not rows:
        if first_study == "Morning" or first_study == "Hosted":
            messagebox.showwarning("Warning", f"No attendance data found for filters: Major={first_major}, Stage={first_stage}, Study=Morning/Hosted, Group={first_group}.")
        else:
            messagebox.showwarning("Warning", f"No attendance data found for filters: Major={first_major}, Stage={first_stage}, Study={first_study}, Group={first_group}.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Major", "Stage", "Study", "Group", "Timestamp", "Attended"])
    for row in rows:
        ws.append(row)
    
    # Create a more descriptive filename for Morning/Hosted exports
    if first_study == "Morning" or first_study == "Hosted":
        study_label = "Morning-Hosted"
    else:
        study_label = first_study
        
    export_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        initialfile=f"{datetime.now().strftime('%Y-%m-%d')}_{first_major}_{first_stage}_{study_label}_{first_group}"
    )
    if export_file:
        wb.save(export_file)
        messagebox.showinfo("Success", f"Attendance exported to {export_file}.")

def reset_attendance():
    """Reset all attendance data and clear the filters."""
    global first_major, first_stage, first_study, first_group
    if not conn:
        messagebox.showerror("Error", "No database loaded.")
        return
    # Clear all attendance records
    cursor = conn.cursor()
    cursor.execute("DELETE FROM attendance")
    conn.commit()
    # Clear the dashboard
    dashboard_tree.delete(*dashboard_tree.get_children())
    # Reset the filters
    first_major = None
    first_stage = None
    first_study = None
    first_group = None
    messagebox.showinfo("Success", "Attendance data and filters have been reset.")

# GUI Setup
root = tk.Tk()
root.title("📊 Attendance Management System")
root.geometry("1200x930")
root.configure(bg="#f0f4f8")
root.resizable(True, True)

# Modern color scheme
PRIMARY_COLOR = "#1a73e8"  # Main blue color
SECONDARY_COLOR = "#4285f4"  # Secondary blue
ACCENT_COLOR = "#0f9d58"  # Green accent
WARNING_COLOR = "#ea4335"  # Red for warnings/errors
DARK_TEXT = "#202124"  # Dark text
LIGHT_TEXT = "#f1f3f4"  # Light text
SURFACE_COLOR = "#ffffff"  # White surface
BORDER_COLOR = "#dadce0"  # Light gray for borders
HOVER_COLOR = "#e8f0fe"  # Light blue for hover states

# Enhanced styling
style = ttk.Style()
style.theme_use('clam')  # Use clam theme as base

# Configure TLabel styles
style.configure("TLabel", font=("Segoe UI", 11), padding=5, background=SURFACE_COLOR)
style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"), foreground=PRIMARY_COLOR, background=SURFACE_COLOR, padding=10)
style.configure("Header.TLabel", font=("Segoe UI", 24, "bold"), foreground=LIGHT_TEXT, background=PRIMARY_COLOR, padding=15)
style.configure("Section.TLabel", font=("Segoe UI", 14, "bold"), foreground=DARK_TEXT, background=SURFACE_COLOR, padding=8)
style.configure("Dashboard.TLabel", font=("Segoe UI", 16, "bold"), foreground=PRIMARY_COLOR, background=SURFACE_COLOR, padding=10)

# Configure button styles
style.configure("TButton", font=("Segoe UI", 11), padding=8)
style.map("TButton", background=[("active", HOVER_COLOR), ("pressed", SECONDARY_COLOR)])

style.configure("Primary.TButton", font=("Segoe UI", 11, "bold"), background=PRIMARY_COLOR, foreground=LIGHT_TEXT)
style.map("Primary.TButton", background=[("active", SECONDARY_COLOR), ("pressed", PRIMARY_COLOR)])

style.configure("Success.TButton", font=("Segoe UI", 11, "bold"), background=ACCENT_COLOR, foreground=LIGHT_TEXT)
style.map("Success.TButton", background=[("active", "#07875e"), ("pressed", ACCENT_COLOR)])

style.configure("Danger.TButton", font=("Segoe UI", 11, "bold"), background=WARNING_COLOR, foreground=LIGHT_TEXT)
style.map("Danger.TButton", background=[("active", "#c62828"), ("pressed", WARNING_COLOR)])

# Configure Treeview
style.configure("Treeview", font=("Segoe UI", 11), rowheight=30, background=SURFACE_COLOR, fieldbackground=SURFACE_COLOR)
style.configure("Treeview.Heading", font=("Segoe UI", 12, "bold"), background=PRIMARY_COLOR, foreground=LIGHT_TEXT)
style.map("Treeview", background=[("selected", HOVER_COLOR)], foreground=[("selected", PRIMARY_COLOR)])

# Custom frame class with rounded corners
class RoundedFrame(tk.Frame):
    def __init__(self, parent, bg_color=SURFACE_COLOR, corner_radius=10, **kwargs):
        super().__init__(parent, **kwargs)
        self.corner_radius = corner_radius
        self.bg_color = bg_color
        
        # Main container
        self.container = tk.Frame(self, bg=bg_color, highlightbackground=BORDER_COLOR, 
                                highlightthickness=1, bd=0)
        self.container.pack(fill="both", expand=True, padx=2, pady=2)

# Create main container with padding
main_container = tk.Frame(root, bg="#f0f4f8", padx=20, pady=20)
main_container.pack(fill="both", expand=True)

# Create header with gradient effect
header_frame = tk.Frame(main_container, bg=PRIMARY_COLOR, height=70)
header_frame.pack(fill="x", pady=(0, 20))

header_label = tk.Label(header_frame, text="Attendance Management System", 
                         font=("Segoe UI", 22, "bold"), fg=LIGHT_TEXT, bg=PRIMARY_COLOR)
header_label.pack(fill="both", expand=True, pady=15)

# Create content area with two columns
content_frame = tk.Frame(main_container, bg="#f0f4f8")
content_frame.pack(fill="both", expand=True)

# Left panel (40% width)
left_panel = RoundedFrame(content_frame, bg_color=SURFACE_COLOR)
left_panel.pack(side=tk.LEFT, fill="both", expand=False, padx=(0, 10), pady=10)

# Right panel (60% width)
right_panel = RoundedFrame(content_frame, bg_color=SURFACE_COLOR)
right_panel.pack(side=tk.RIGHT, fill="both", expand=True, padx=(10, 0), pady=10)

# Left panel content
left_content = tk.Frame(left_panel.container, bg=SURFACE_COLOR)
left_content.pack(fill="both", expand=True, padx=15, pady=15)

# Logo section
logo_frame = tk.Frame(left_content, bg=SURFACE_COLOR)
logo_frame.pack(fill="x", pady=(0, 20))

try:
    university_logo = Image.open(resource_path("Any_logo.png")).resize((200, 220))
    university_logo_img = ImageTk.PhotoImage(university_logo)
    logo_label = tk.Label(logo_frame, image=university_logo_img, bg=SURFACE_COLOR)
    logo_label.image = university_logo_img
    logo_label.pack(pady=10)
except Exception as e:
    print("Logo not found:", e)
    logo_label = tk.Label(logo_frame, text="University Logo", font=("Segoe UI", 16), bg=SURFACE_COLOR)
    logo_label.pack(pady=10)

# Database section
db_section = tk.Frame(left_content, bg=SURFACE_COLOR)
db_section.pack(fill="x", pady=15)

db_title = tk.Label(db_section, text="Database Management", font=("Segoe UI", 14, "bold"), 
                    fg=PRIMARY_COLOR, bg=SURFACE_COLOR)
db_title.pack(anchor="w", pady=(0, 10))

db_frame = tk.Frame(db_section, bg=SURFACE_COLOR)
db_frame.pack(fill="x")

db_label = tk.Label(db_frame, text="Select Database:", font=("Segoe UI", 11), bg=SURFACE_COLOR)
db_label.pack(anchor="w", pady=(5, 8))

# Create a frame for dropdown with border
dropdown_container = tk.Frame(db_frame, bg=SURFACE_COLOR, highlightbackground=BORDER_COLOR, 
                           highlightthickness=1, bd=0)
dropdown_container.pack(fill="x", pady=(0, 10))

db_dropdown = ttk.Combobox(dropdown_container, state="readonly", font=("Segoe UI", 11), width=25)
db_dropdown.pack(fill="x", padx=2, pady=2)
populate_db_dropdown()
db_dropdown.bind("<<ComboboxSelected>>", load_db_from_dropdown)

# Create new database button
btn_create_db = ttk.Button(db_frame, text="Create New Database", command=create_db, style="Primary.TButton")
btn_create_db.pack(fill="x", pady=10)

# Ministry logo or second logo
try:
    ministry_logo = Image.open(resource_path("Any_logo2.png")).resize((200, 180))
    ministry_logo_img = ImageTk.PhotoImage(ministry_logo)
    ministry_label = tk.Label(left_content, image=ministry_logo_img, bg=SURFACE_COLOR)
    ministry_label.image = ministry_logo_img
    ministry_label.pack(pady=15)
except Exception as e:
    print("Secondary logo not found:", e)

# Credit label at bottom
credit_frame = tk.Frame(left_content, bg=SURFACE_COLOR)
credit_frame.pack(side=tk.BOTTOM, fill="x", pady=15)

credit_label = tk.Label(credit_frame, text="Developed by Ahmad Tchnology", 
                       font=("Segoe UI", 10, "italic"), fg="#666", bg=SURFACE_COLOR)
credit_label.pack(side=tk.BOTTOM)

# Right panel content with notebook (tabs)
right_content = tk.Frame(right_panel.container, bg=SURFACE_COLOR)
right_content.pack(fill="both", expand=True, padx=15, pady=15)

# Create a notebook (tabbed interface)
notebook = ttk.Notebook(right_content)
notebook.pack(fill="both", expand=True)

# Style the notebook
style.configure("TNotebook", background=SURFACE_COLOR)
style.configure("TNotebook.Tab", font=("Segoe UI", 11), padding=[15, 5])
style.map("TNotebook.Tab", background=[("selected", PRIMARY_COLOR)], 
         foreground=[("selected", LIGHT_TEXT)])

# Create tabs
student_tab = tk.Frame(notebook, bg=SURFACE_COLOR)
attendance_tab = tk.Frame(notebook, bg=SURFACE_COLOR)

notebook.add(student_tab, text="Student Management")
notebook.add(attendance_tab, text="Attendance Recording")

# ====== Student Management Tab ======
student_frame = tk.Frame(student_tab, bg=SURFACE_COLOR, padx=15, pady=15)
student_frame.pack(fill="both", expand=True)

student_title = tk.Label(student_frame, text="Add New Student", font=("Segoe UI", 16, "bold"), 
                        fg=PRIMARY_COLOR, bg=SURFACE_COLOR)
student_title.pack(anchor="w", pady=(0, 15))

# Create a frame for the student form
form_frame = tk.Frame(student_frame, bg=SURFACE_COLOR)
form_frame.pack(fill="both", pady=10)

# Helper function to create form fields
def create_form_field(parent, row, label_text, widget):
    field_frame = tk.Frame(parent, bg=SURFACE_COLOR)
    field_frame.pack(fill="x", pady=8)
    
    label = tk.Label(field_frame, text=label_text, font=("Segoe UI", 11), 
                    width=15, anchor='w', bg=SURFACE_COLOR)
    label.pack(side=tk.LEFT)
    
    input_container = tk.Frame(field_frame, bg=SURFACE_COLOR, highlightbackground=BORDER_COLOR, 
                             highlightthickness=1, bd=0)
    input_container.pack(side=tk.LEFT, fill="x", expand=True)
    
    widget.configure(font=("Segoe UI", 11))
    widget.pack(fill="x", padx=2, pady=2)
    
    return widget

# Student ID
entry_id = ttk.Entry(form_frame)
create_form_field(form_frame, 0, "Student ID:", entry_id)

# Name
entry_name = ttk.Entry(form_frame)
create_form_field(form_frame, 1, "Name:", entry_name)

# Major
entry_major = ttk.Entry(form_frame)
create_form_field(form_frame, 2, "Major:", entry_major)

# Stage
entry_stage = ttk.Entry(form_frame)
create_form_field(form_frame, 3, "Stage:", entry_stage)

# Study
entry_study = ttk.Combobox(form_frame, values=["Morning", "Evening", "Hosted"])
create_form_field(form_frame, 4, "Study:", entry_study)
entry_study.current(0)

# Group
entry_group = ttk.Entry(form_frame)
create_form_field(form_frame, 5, "Group:", entry_group)

# Buttons Frame
buttons_frame = tk.Frame(form_frame, bg=SURFACE_COLOR)
buttons_frame.pack(fill="x", pady=15)

# Add Student Button
btn_add_student = ttk.Button(buttons_frame, text="Add Student", command=add_student, style="Primary.TButton")
btn_add_student.pack(side=tk.LEFT, padx=(0, 10))

# Import Students Button
btn_import = ttk.Button(buttons_frame, text="Import from Excel", command=import_students_from_excel, style="Success.TButton")
btn_import.pack(side=tk.LEFT)

# ====== Attendance Tab ======
attendance_frame = tk.Frame(attendance_tab, bg=SURFACE_COLOR, padx=15, pady=15)
attendance_frame.pack(fill="both", expand=True)

attendance_title = tk.Label(attendance_frame, text="NFC Attendance Recording", font=("Segoe UI", 16, "bold"), 
                          fg=PRIMARY_COLOR, bg=SURFACE_COLOR)
attendance_title.pack(anchor="w", pady=(0, 15))

# NFC Input section
nfc_frame = tk.Frame(attendance_frame, bg=SURFACE_COLOR)
nfc_frame.pack(fill="x", pady=10)

nfc_label = tk.Label(nfc_frame, text="Scan NFC Card:", font=("Segoe UI", 11), bg=SURFACE_COLOR)
nfc_label.pack(anchor="w", pady=(5, 8))

# Highlighted NFC input field
nfc_container = tk.Frame(nfc_frame, bg=SURFACE_COLOR, highlightbackground=PRIMARY_COLOR, 
                        highlightthickness=2, bd=0)
nfc_container.pack(fill="x", pady=(0, 15))

entry_nfc = ttk.Entry(nfc_container, font=("Segoe UI", 12, "bold"), width=30)
entry_nfc.pack(fill="x", padx=3, pady=3)
entry_nfc.bind("<Return>", record_attendance)
entry_nfc.focus()  # Set focus to the NFC entry field

# Attendance action buttons
attendance_buttons = tk.Frame(attendance_frame, bg=SURFACE_COLOR)
attendance_buttons.pack(fill="x", pady=15)

btn_record = ttk.Button(attendance_buttons, text="📝 Record Attendance", 
                       command=record_attendance, style="Primary.TButton")
btn_record.pack(side=tk.LEFT, padx=(0, 10))

btn_export = ttk.Button(attendance_buttons, text="📤 Export to Excel", 
                       command=export_attendance, style="Success.TButton")
btn_export.pack(side=tk.LEFT, padx=(0, 10))

btn_reset = ttk.Button(attendance_buttons, text="🔄 Reset Attendance", 
                      command=reset_attendance, style="Danger.TButton")
btn_reset.pack(side=tk.LEFT)

# Attendance dashboard
dashboard_frame = tk.Frame(attendance_frame, bg=SURFACE_COLOR)
dashboard_frame.pack(fill="both", expand=True, pady=15)

dashboard_label = tk.Label(dashboard_frame, text="Attendance Dashboard", 
                          font=("Segoe UI", 14, "bold"), fg=PRIMARY_COLOR, bg=SURFACE_COLOR)
dashboard_label.pack(anchor="w", pady=(0, 10))

# Treeview container with border
tree_container = tk.Frame(dashboard_frame, bg=SURFACE_COLOR, highlightbackground=BORDER_COLOR, 
                         highlightthickness=1, bd=0)
tree_container.pack(fill="both", expand=True)

# Create Treeview
dashboard_tree = ttk.Treeview(tree_container, columns=("Name", "Major", "Stage", "Study", "Group", "Timestamp", "Attended"), 
                             show="headings", height=10)

# Set column headings
dashboard_tree.heading("Name", text="Name")
dashboard_tree.heading("Major", text="Major")
dashboard_tree.heading("Stage", text="Stage")
dashboard_tree.heading("Study", text="Study")
dashboard_tree.heading("Group", text="Group")
dashboard_tree.heading("Timestamp", text="Timestamp")
dashboard_tree.heading("Attended", text="Attended")

# Set column widths
dashboard_tree.column("Name", width=150)
dashboard_tree.column("Major", width=120)
dashboard_tree.column("Stage", width=80)
dashboard_tree.column("Study", width=100)
dashboard_tree.column("Group", width=80)
dashboard_tree.column("Timestamp", width=140)
dashboard_tree.column("Attended", width=80)

# Pack the treeview with scrollbar
dashboard_tree.pack(side=tk.LEFT, fill="both", expand=True, padx=1, pady=1)

# Add scrollbar
scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=dashboard_tree.yview)
dashboard_tree.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill="y")

# Run the application
root.mainloop()